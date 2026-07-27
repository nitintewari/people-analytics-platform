"""
Polished Excel report — closes Advanced Excel gap from JD
3 sheets: Executive Summary, Dept Breakdown, High Risk (HR only)
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

df   = pd.read_csv("output/attrition_scored.csv")
dept = pd.read_csv("output/anomaly_flags.csv")

wb = Workbook()
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="1F4E79")

def style_headers(ws, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

# Sheet 1: Executive Summary
ws1 = wb.active
ws1.title = "Executive Summary"
style_headers(ws1, ["Metric", "Value"])

metrics = [
    ("Total Headcount",            len(df)),
    ("Attrition Rate",             f"{(df['Attrition']=='Yes').mean():.1%}"),
    ("High Risk Employees",        int((df['RiskSegment']=='High').sum())),
    ("Avg Engagement Score (1-4)", f"{df['EngagementScore'].mean():.2f}"),
    ("Avg Annual Salary",          f"${df['AnnualSalary'].mean():,.0f}"),
    ("Anomalous Departments",      int(dept['AnomalyFlag'].sum())),
]
for row, (m, v) in enumerate(metrics, 2):
    ws1.cell(row=row, column=1, value=m)
    ws1.cell(row=row, column=2, value=v)
ws1.column_dimensions["A"].width = 30
ws1.column_dimensions["B"].width = 20

# Sheet 2: Department Breakdown
ws2 = wb.create_sheet("Department Breakdown")
ds = (df.groupby("Department").agg(
    Headcount=("EmployeeID","count"),
    AttritionRate=("Attrition", lambda s: (s=="Yes").mean()),
    AvgSalary=("AnnualSalary","mean"),
    AvgEngagement=("EngagementScore","mean"),
    HighRiskCount=("RiskSegment", lambda s: (s=="High").sum()),
).reset_index().round(3))
ds["AttritionRate"] = ds["AttritionRate"].map("{:.1%}".format)
ds["AvgSalary"]     = ds["AvgSalary"].map("${:,.0f}".format)

style_headers(ws2, list(ds.columns))
for row_data in dataframe_to_rows(ds, index=False, header=False):
    ws2.append(row_data)
for col in ws2.columns:
    ws2.column_dimensions[col[0].column_letter].width = 20

# Sheet 3: High Risk — HR Partners only
ws3 = wb.create_sheet("High Risk - HR Only")
hr_df = df[df["RiskSegment"]=="High"][[
    "EmployeeID","Department","JobRole","TenureYears","AnnualSalary",
    "JobSatisfaction","EngagementScore","OverTime","AttritionRiskScore"
]].sort_values("AttritionRiskScore", ascending=False)

style_headers(ws3, list(hr_df.columns))
for row_data in dataframe_to_rows(hr_df, index=False, header=False):
    ws3.append(row_data)
for col in ws3.columns:
    ws3.column_dimensions[col[0].column_letter].width = 20

wb.save("output/people_analytics_executive_report.xlsx")
print("Excel saved: output/people_analytics_executive_report.xlsx")
print(f"  Sheet 1: {len(metrics)} metrics")
print(f"  Sheet 2: {len(ds)} departments")
print(f"  Sheet 3: {len(hr_df)} high-risk employees")